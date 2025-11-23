import openpyxl
from openpyxl import Workbook

read_wb = openpyxl.load_workbook("input_data.xlsx")
read_sheet = read_wb.active

write_wb = Workbook()
write_sheet = write_wb.active

for i in range(1, read_sheet.max_row + 1):
    for j in range(1, read_sheet.max_column + 1):
        cell_value = read_sheet.cell(row=i, column=j).value
        write_sheet.cell(row=i, column=j).value = cell_value

write_wb.save("output_data.xlsx")
